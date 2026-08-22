"""
XLSX -> SQLite ingestion, and read-side query helpers used by tools.

Design notes (see docs/ARCHITECTURE.md for full rationale):
- Pandas is used ONLY during ingestion, never as the live query engine.
- All runtime reads use parameterized SQL through small, named functions —
  never raw/arbitrary SQL built from user or model input.
- Column names below match the real workbook headers exactly:
    accounts: account_id, account_name, plan, status, csm, contract_file,
              premium_support, notes
    orders:   order_id, account_id, carrier, status, booked_at,
              pickup_window_start, pickup_window_end, pickup_actual_at,
              shipment_fee_inr, carrier_fault, customer_fault,
              cancellation_requested_at, notes
    tickets:  ticket_id, account_id, created_at, status, subject,
              description, channel, assigned_to, last_customer_message_at,
              historical_resolution
"""
import sqlite3
from contextlib import contextmanager

import openpyxl


SCHEMA = """
CREATE TABLE accounts (
    account_id TEXT PRIMARY KEY,
    account_name TEXT NOT NULL,
    plan TEXT NOT NULL,
    status TEXT NOT NULL,
    csm TEXT,
    contract_file TEXT,
    premium_support INTEGER NOT NULL DEFAULT 0,
    notes TEXT
);

CREATE TABLE orders (
    order_id TEXT PRIMARY KEY,
    account_id TEXT NOT NULL REFERENCES accounts(account_id),
    carrier TEXT,
    status TEXT NOT NULL,
    booked_at TEXT,
    pickup_window_start TEXT,
    pickup_window_end TEXT,
    pickup_actual_at TEXT,
    shipment_fee_inr REAL,
    carrier_fault INTEGER NOT NULL DEFAULT 0,
    customer_fault INTEGER NOT NULL DEFAULT 0,
    cancellation_requested_at TEXT,
    notes TEXT
);

CREATE TABLE tickets (
    ticket_id TEXT PRIMARY KEY,
    account_id TEXT NOT NULL REFERENCES accounts(account_id),
    created_at TEXT,
    status TEXT NOT NULL,
    subject TEXT,
    description TEXT,
    channel TEXT,
    assigned_to TEXT,
    last_customer_message_at TEXT,
    historical_resolution TEXT
);

CREATE INDEX idx_orders_account ON orders(account_id);
CREATE INDEX idx_tickets_account ON tickets(account_id);
"""


def _bool_to_int(v):
    return 1 if bool(v) else 0


def ingest_xlsx_to_sqlite(xlsx_path: str, db_path: str) -> dict:
    wb = openpyxl.load_workbook(xlsx_path, data_only=True)

    conn = sqlite3.connect(db_path)
    cur = conn.cursor()
    cur.executescript("""
        DROP TABLE IF EXISTS accounts;
        DROP TABLE IF EXISTS orders;
        DROP TABLE IF EXISTS tickets;
    """)
    cur.executescript(SCHEMA)

    summary = {}

    # --- accounts ---
    ws = wb["accounts"]
    rows = list(ws.iter_rows(values_only=True))
    header, data_rows = rows[0], rows[1:]
    assert header == (
        "account_id", "account_name", "plan", "status", "csm",
        "contract_file", "premium_support", "notes",
    ), f"Unexpected accounts header: {header}"
    for r in data_rows:
        cur.execute(
            "INSERT INTO accounts VALUES (?,?,?,?,?,?,?,?)",
            (r[0], r[1], r[2], r[3], r[4], r[5], _bool_to_int(r[6]), r[7]),
        )
    summary["accounts"] = len(data_rows)

    # --- orders ---
    ws = wb["orders"]
    rows = list(ws.iter_rows(values_only=True))
    header, data_rows = rows[0], rows[1:]
    assert header == (
        "order_id", "account_id", "carrier", "status", "booked_at",
        "pickup_window_start", "pickup_window_end", "pickup_actual_at",
        "shipment_fee_inr", "carrier_fault", "customer_fault",
        "cancellation_requested_at", "notes",
    ), f"Unexpected orders header: {header}"
    for r in data_rows:
        r = list(r)
        # normalize datetimes to isoformat strings
        for idx in (4, 5, 6, 7, 11):
            if hasattr(r[idx], "isoformat"):
                r[idx] = r[idx].isoformat(sep=" ")
        cur.execute(
            "INSERT INTO orders VALUES (?,?,?,?,?,?,?,?,?,?,?,?,?)",
            (
                r[0], r[1], r[2], r[3], r[4], r[5], r[6], r[7],
                r[8], _bool_to_int(r[9]), _bool_to_int(r[10]), r[11], r[12],
            ),
        )
    summary["orders"] = len(data_rows)

    # --- tickets ---
    ws = wb["tickets"]
    rows = list(ws.iter_rows(values_only=True))
    header, data_rows = rows[0], rows[1:]
    assert header == (
        "ticket_id", "account_id", "created_at", "status", "subject",
        "description", "channel", "assigned_to",
        "last_customer_message_at", "historical_resolution",
    ), f"Unexpected tickets header: {header}"
    for r in data_rows:
        r = list(r)
        for idx in (2, 8):
            if hasattr(r[idx], "isoformat"):
                r[idx] = r[idx].isoformat(sep=" ")
        cur.execute(
            "INSERT INTO tickets VALUES (?,?,?,?,?,?,?,?,?,?)",
            tuple(r),
        )
    summary["tickets"] = len(data_rows)

    conn.commit()
    conn.close()
    return summary


@contextmanager
def get_conn(db_path: str):
    conn = sqlite3.connect(db_path)
    conn.row_factory = sqlite3.Row
    try:
        yield conn
    finally:
        conn.close()


# ---------------------------------------------------------------------------
# Bounded read operations. These are the ONLY way tools may reach the DB.
# Every function takes explicit args, uses parameterized SQL, and returns
# plain dicts. No function here accepts a raw SQL string.
# ---------------------------------------------------------------------------

def get_account(db_path: str, account_id: str) -> dict | None:
    with get_conn(db_path) as conn:
        row = conn.execute(
            "SELECT * FROM accounts WHERE account_id = ?", (account_id,)
        ).fetchone()
        return dict(row) if row else None


def get_order(db_path: str, order_id: str) -> dict | None:
    with get_conn(db_path) as conn:
        row = conn.execute(
            "SELECT * FROM orders WHERE order_id = ?", (order_id,)
        ).fetchone()
        return dict(row) if row else None


def get_ticket(db_path: str, ticket_id: str) -> dict | None:
    with get_conn(db_path) as conn:
        row = conn.execute(
            "SELECT * FROM tickets WHERE ticket_id = ?", (ticket_id,)
        ).fetchone()
        return dict(row) if row else None


def list_account_orders(db_path: str, account_id: str) -> list[dict]:
    with get_conn(db_path) as conn:
        rows = conn.execute(
            "SELECT * FROM orders WHERE account_id = ? ORDER BY booked_at",
            (account_id,),
        ).fetchall()
        return [dict(r) for r in rows]


def list_account_tickets(db_path: str, account_id: str) -> list[dict]:
    with get_conn(db_path) as conn:
        rows = conn.execute(
            "SELECT * FROM tickets WHERE account_id = ? ORDER BY created_at",
            (account_id,),
        ).fetchall()
        return [dict(r) for r in rows]


def find_tickets_by_keyword(db_path: str, keyword: str, account_id: str | None = None) -> list[dict]:
    like = f"%{keyword}%"
    with get_conn(db_path) as conn:
        if account_id:
            rows = conn.execute(
                """SELECT * FROM tickets
                   WHERE account_id = ? AND (subject LIKE ? OR description LIKE ?)
                   ORDER BY created_at""",
                (account_id, like, like),
            ).fetchall()
        else:
            rows = conn.execute(
                """SELECT * FROM tickets
                   WHERE subject LIKE ? OR description LIKE ?
                   ORDER BY created_at""",
                (like, like),
            ).fetchall()
        return [dict(r) for r in rows]


def find_open_tickets(db_path: str, account_id: str | None = None) -> list[dict]:
    with get_conn(db_path) as conn:
        if account_id:
            rows = conn.execute(
                "SELECT * FROM tickets WHERE status = 'open' AND account_id = ? ORDER BY created_at",
                (account_id,),
            ).fetchall()
        else:
            rows = conn.execute(
                "SELECT * FROM tickets WHERE status = 'open' ORDER BY created_at"
            ).fetchall()
        return [dict(r) for r in rows]


def list_all_orders(db_path: str) -> list[dict]:
    """Internal/ops-only aggregate view — callers must check role before using this."""
    with get_conn(db_path) as conn:
        rows = conn.execute("SELECT * FROM orders").fetchall()
        return [dict(r) for r in rows]


def list_all_open_tickets(db_path: str) -> list[dict]:
    """Internal/ops-only aggregate view — callers must check role before using this."""
    with get_conn(db_path) as conn:
        rows = conn.execute("SELECT * FROM tickets WHERE status = 'open'").fetchall()
        return [dict(r) for r in rows]
